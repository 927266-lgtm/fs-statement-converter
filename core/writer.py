import os
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def write_outputs(chunks: list, registry: list, output_dir: str, file_path: str) -> dict:
    """
    Writes all parsed chunks and registry data to the output directory.
    Generates:
      - /chunks/<chunk_id>.json
      - chunk_registry.xlsx
      - extraction_report.xlsx
      - registry.json
      - <doc_name>_full_text.txt
    """
    os.makedirs(output_dir, exist_ok=True)
    chunks_dir = os.path.join(output_dir, "chunks")
    os.makedirs(chunks_dir, exist_ok=True)
    
    doc_stem = os.path.basename(file_path).rsplit(".", 1)[0]
    
    # 1. Write individual JSON chunk files
    for c in chunks:
        chunk_file = os.path.join(chunks_dir, f"{c['chunk_id']}.json")
        with open(chunk_file, "w", encoding="utf-8") as f:
            json.dump(c, f, indent=2, ensure_ascii=False)
            
    # 2. Write machine registry backup JSON
    reg_json_file = os.path.join(output_dir, "registry.json")
    with open(reg_json_file, "w", encoding="utf-8") as f:
        json.dump(registry, f, indent=2, ensure_ascii=False)
        
    # 3. Write full text compilation file for LLM upload
    full_text_file = os.path.join(output_dir, f"{doc_stem}_full_text.txt")
    with open(full_text_file, "w", encoding="utf-8") as f:
        f.write(f"======================================================================\n")
        f.write(f"COMPILATION OF EXTRACTED FINANCIAL STATEMENTS & SINGLE AUDIT REPORT\n")
        f.write(f"Source Document: {os.path.basename(file_path)}\n")
        f.write(f"======================================================================\n\n")
        
        for c in chunks:
            f.write(f"\n----- START CHUNK: {c['label']} (ID: {c['chunk_id']}) -----\n")
            f.write(c["content"])
            f.write(f"\n----- END CHUNK: {c['label']} -----\n")
            
    # 4. Write Excel Chunk Registry (chunk_registry.xlsx)
    wb_reg = openpyxl.Workbook()
    ws_reg = wb_reg.active
    ws_reg.title = "Chunk Registry"
    ws_reg.views.sheetView[0].showGridLines = True
    
    # Header styling
    header_fill = PatternFill(start_color="1A3A5C", end_color="1A3A5C", fill_type="solid") # Dark Blue Accent
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    headers = [
        "Chunk ID", "Label", "Section Type", "Start Page", "End Page",
        "Char Count", "Word Count", "Quality Level", "OCR Needed", "Audit Math Verification"
    ]
    
    for col_idx, h in enumerate(headers, 1):
        cell = ws_reg.cell(row=1, column=col_idx)
        cell.value = h
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    # Content rows
    thin_border = Border(
        left=Side(style='thin', color='D4CFC8'),
        right=Side(style='thin', color='D4CFC8'),
        top=Side(style='thin', color='D4CFC8'),
        bottom=Side(style='thin', color='D4CFC8')
    )
    
    # Fills for Quality
    high_fill = PatternFill(start_color="D8F3DC", end_color="D8F3DC", fill_type="solid") # Light green
    med_fill = PatternFill(start_color="FFF3CD", end_color="FFF3CD", fill_type="solid") # Light yellow
    low_fill = PatternFill(start_color="FDE8E8", end_color="FDE8E8", fill_type="solid") # Light red
    
    for row_idx, r in enumerate(registry, 2):
        ws_reg.cell(row=row_idx, column=1, value=r["chunk_id"]).alignment = align_left
        ws_reg.cell(row=row_idx, column=2, value=r["label"]).alignment = align_left
        ws_reg.cell(row=row_idx, column=3, value=r["type"]).alignment = align_center
        ws_reg.cell(row=row_idx, column=4, value=r["start_page"]).alignment = align_center
        ws_reg.cell(row=row_idx, column=5, value=r["end_page"]).alignment = align_center
        ws_reg.cell(row=row_idx, column=6, value=r["char_count"]).alignment = align_center
        ws_reg.cell(row=row_idx, column=7, value=r["word_count"]).alignment = align_center
        
        # Quality cell with fill
        q_cell = ws_reg.cell(row=row_idx, column=8, value=r["quality"])
        q_cell.alignment = align_center
        if r["quality"] == "HIGH":
            q_cell.fill = high_fill
            q_cell.font = Font(name="Calibri", size=10, bold=True, color="2D6A4F")
        elif r["quality"] == "MED":
            q_cell.fill = med_fill
            q_cell.font = Font(name="Calibri", size=10, bold=True, color="B5580A")
        else:
            q_cell.fill = low_fill
            q_cell.font = Font(name="Calibri", size=10, bold=True, color="8B1A1A")
            
        ocr_cell = ws_reg.cell(row=row_idx, column=9, value=r["ocr_needed"])
        ocr_cell.alignment = align_center
        if r["ocr_needed"] == "YES":
            ocr_cell.fill = low_fill
            ocr_cell.font = Font(name="Calibri", size=10, bold=True, color="8B1A1A")
            
        ws_reg.cell(row=row_idx, column=10, value=r["balance_check"]).alignment = align_left
        
        # Set border
        for col_idx in range(1, len(headers) + 1):
            ws_reg.cell(row=row_idx, column=col_idx).border = thin_border
            
    # Autofit columns
    for col in ws_reg.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                # Limit col width to max 40 for readability
                max_len = max(max_len, min(len(str(cell.value)), 45))
        col_letter = get_column_letter(col[0].column)
        ws_reg.column_dimensions[col_letter].width = max_len + 3
        
    wb_reg.save(os.path.join(output_dir, "chunk_registry.xlsx"))
    
    # 5. Write Excel Extraction Report (extraction_report.xlsx)
    wb_rep = openpyxl.Workbook()
    ws_rep = wb_rep.active
    ws_rep.title = "Completeness Report"
    ws_rep.views.sheetView[0].showGridLines = True
    
    # Write summary stats
    total_chunks = len(chunks)
    high_q = sum(1 for c in chunks if c["quality"] == "HIGH")
    med_q = sum(1 for c in chunks if c["quality"] == "MED")
    low_q = sum(1 for c in chunks if c["quality"] == "LOW")
    ocr_needed = sum(1 for c in chunks if c["is_scanned"])
    
    # Find balance checks result
    balance_status = "Not Detected"
    for c in chunks:
        if c.get("audit_math", {}).get("balanced") is True:
            balance_status = "PASSED (Balance Sheet Equations match)"
            break
        elif c.get("audit_math", {}).get("balanced") is False:
            balance_status = "FAILED (Balance Sheet equation mismatch)"
            break
            
    # Title row
    ws_rep.cell(row=1, column=1, value="Extraction & Completeness Report").font = Font(name="Calibri", size=16, bold=True, color="1A3A5C")
    ws_rep.cell(row=2, column=1, value=f"Document: {os.path.basename(file_path)}").font = Font(name="Calibri", size=11, italic=True)
    
    # KPI metrics table
    metrics_headers = ["Metric Parameter", "Value", "Status / Notes"]
    for col_idx, h in enumerate(metrics_headers, 1):
        cell = ws_rep.cell(row=4, column=col_idx, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = align_center
        
    kpis = [
        ("Total Logical Chunks", total_chunks, "Parsed from report sections"),
        ("Quality: HIGH Chunks", high_q, "Audited financial tables"),
        ("Quality: MED Chunks", med_q, "Standard readable text sections"),
        ("Quality: LOW Chunks", low_q, "OCR required or empty sections"),
        ("Scanned Pages (OCR recommended)", ocr_needed, "No native text detected"),
        ("Balance Sheet Audit Math", balance_status, "Verification check: Assets = Liabilities + Equity")
    ]
    
    for row_offset, kpi in enumerate(kpis, 5):
        m, v, s = kpi
        ws_rep.cell(row=row_offset, column=1, value=m).alignment = align_left
        ws_rep.cell(row=row_offset, column=2, value=v).alignment = align_center
        ws_rep.cell(row=row_offset, column=3, value=s).alignment = align_left
        
        # Border and Styling
        for c_idx in range(1, 4):
            c = ws_rep.cell(row=row_offset, column=c_idx)
            c.border = thin_border
            # Conditional coloring
            if m == "Balance Sheet Audit Math":
                if "PASSED" in str(v):
                    c.fill = high_fill
                    c.font = Font(name="Calibri", size=11, bold=True, color="2D6A4F")
                elif "FAILED" in str(v):
                    c.fill = low_fill
                    c.font = Font(name="Calibri", size=11, bold=True, color="8B1A1A")
            elif m == "Scanned Pages (OCR recommended)" and v > 0:
                c.fill = low_fill
                
    ws_rep.column_dimensions['A'].width = 35
    ws_rep.column_dimensions['B'].width = 35
    ws_rep.column_dimensions['C'].width = 50
    
    wb_rep.save(os.path.join(output_dir, "extraction_report.xlsx"))
    
    return {
        "total_chunks": total_chunks,
        "quality_high": high_q,
        "quality_med": med_q,
        "quality_low": low_q,
        "ocr_chunks": ocr_needed,
        "output_dir": output_dir
    }
