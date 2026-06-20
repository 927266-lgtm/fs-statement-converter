import os
import openpyxl

def extract_excel(file_path: str) -> dict:
    """
    Extracts text and table information from a financial statement Excel file.
    Treats each sheet as a separate 'page'.
    """
    raw_doc = {
        "filename": os.path.basename(file_path),
        "pages": [],
        "total_pages": 0,
        "ocr_needed_pages": 0
    }
    
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True, read_only=True)
        raw_doc["total_pages"] = len(wb.sheetnames)
        
        for idx, sname in enumerate(wb.sheetnames):
            page_num = idx + 1
            sheet = wb[sname]
            
            # Extract rows
            rows = []
            max_cols = 0
            for r in sheet.iter_rows(values_only=True):
                # Trim trailing Nones to clean up the row
                row_data = list(r)
                while row_data and row_data[-1] is None:
                    row_data.pop()
                if row_data:
                    rows.append(row_data)
                    max_cols = max(max_cols, len(row_data))
                    
            # Skip empty sheets
            if not rows:
                continue
                
            # Build grid text representation
            col_widths = [0] * max_cols
            for r in rows:
                for c_idx, cell in enumerate(r):
                    if cell is not None:
                        col_widths[c_idx] = max(col_widths[c_idx], len(str(cell)))
                        
            text_lines = []
            for r in rows:
                line = []
                for c_idx, cell in enumerate(r):
                    cell_str = str(cell) if cell is not None else ""
                    # Pad cell according to column width plus standard spacer
                    line.append(cell_str.ljust(col_widths[c_idx] + 3))
                text_lines.append("".join(line).rstrip())
                
            grid_text = f"Sheet: {sname}\n" + "=" * (len(sname) + 7) + "\n" + "\n".join(text_lines)
            
            raw_doc["pages"].append({
                "page_number": page_num,
                "sheet_name": sname,
                "text": grid_text,
                "raw_text": grid_text,
                "tables": [rows], # Store entire sheet as a table
                "is_scanned": False
            })
            
    except Exception as e:
        raw_doc["error"] = f"Failed to parse Excel: {str(e)}"
        
    return raw_doc
